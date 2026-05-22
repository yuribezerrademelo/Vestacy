import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from files.config import DOWNLOAD_DIR

# ============================================================
# CONFIGURAÇÕES - ajuste apenas esta seção se necessário
# ============================================================

# Pasta onde estão os 8 arquivos exportados da Mtrix
# Importado do config.py — mesma lógica do download_watcher.py
PASTA_BASES = Path(DOWNLOAD_DIR) if DOWNLOAD_DIR else Path.home() / "Downloads"

# Clientes a serem removidos das bases normais
CLIENTES_MATEUS = [
    "ARMAZEM MATEUS LTDA - DB PI",
    "ARMAZEM MATEUS LTDA - DB MA",
]

# Coluna de identificação do cliente
COLUNA_CLIENTE = "Grp. Econômico"

# Pares: base normal -> base Mateus equivalente
PARES_BASES = {
    "01.01 - Base Grit":     "01.01 - Base Grit Mateus",
    "02.01 - Categorias Ar": "02.01 - Categorias Ar MATEUS",
    "03.01 - Actual Dbs":    "03.01 - Actual Dbs Mateus",
    "05.01 - Produtos":      "05.01 - Produtos Mateus",
}

# Linhas a apagar APÓS o cabeçalho (índice 0 = primeira linha de dados)
# Serão apagadas antes de qualquer outro tratamento
LINHAS_EXTRAS = {
    # Bases que têm 1 linha em branco (linha 2 do Excel = índice 0 nos dados)
    "01.01 - Base Grit":             [0],
    "01.01 - Base Grit Mateus":      [0],
    "02.01 - Categorias Ar":         [0],
    "02.01 - Categorias Ar MATEUS":  [0],
    "05.01 - Produtos":              [0],
    "05.01 - Produtos Mateus":       [0],
    # Bases que têm 4 linhas em branco (linhas 2-5 do Excel = índices 0-3 nos dados)
    "03.01 - Actual Dbs":            [0, 1, 2, 3],
    "03.01 - Actual Dbs Mateus":     [0, 1, 2, 3],
}

# ============================================================
# FUNÇÕES
# ============================================================

def caminho(nome_base):
    return PASTA_BASES / f"{nome_base}.xlsx"


def celula_para_texto(cell):
    """
    Retorna o valor da célula como texto preservando máxima precisão.
    - Números: sempre com todas as casas decimais reais (sem arredondamento)
    - Datas: formato DD/MM/AAAA
    - Texto: valor direto
    """
    from openpyxl.styles.numbers import is_date_format

    value = cell.value

    if value is None:
        return ""

    # Data
    fmt = cell.number_format or ""
    if is_date_format(fmt):
        if hasattr(value, "strftime"):
            return value.strftime("%d/%m/%Y")
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(value).strftime("%d/%m/%Y")
        except Exception:
            return str(value)

    # Número — preserva máxima precisão, converte para padrão BR
    if isinstance(value, float):
        # Converte para string com precisão total (sem notação científica)
        texto = f"{value:.15g}"
        # Troca ponto por vírgula (padrão BR)
        return texto.replace(".", ",")

    if isinstance(value, int):
        return str(value)

    # Texto ou qualquer outro tipo
    return str(value)


def carregar_base(nome_base):
    """Carrega o arquivo Excel preservando máxima precisão numérica."""
    arquivo = caminho(nome_base)
    if not arquivo.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {arquivo}")

    wb = load_workbook(arquivo, data_only=True)
    ws = wb.active

    linhas = list(ws.iter_rows())
    if not linhas:
        raise ValueError(f"Arquivo vazio: {arquivo}")

    # Primeira linha = cabeçalho (valor bruto)
    cabecalho = [cell.value if cell.value is not None else "" for cell in linhas[0]]

    # Demais linhas = dados com precisão total
    dados = []
    for row in linhas[1:]:
        dados.append([celula_para_texto(cell) for cell in row])

    df = pd.DataFrame(dados, columns=cabecalho)
    print(f"  Carregado '{nome_base}.xlsx' -> {len(df)} linhas")
    return df


def remover_linhas_extras(df, nome_base):
    """Remove linhas em branco logo abaixo do cabeçalho."""
    indices = LINHAS_EXTRAS.get(nome_base, [])
    if indices:
        indices_validos = [i for i in indices if i < len(df)]
        df = df.drop(index=indices_validos).reset_index(drop=True)
        print(f"  Removidas {len(indices_validos)} linha(s) extras em '{nome_base}'")
    return df


def remover_clientes_mateus(df, nome_base):
    """Remove todas as linhas dos 2 clientes Mateus."""
    if COLUNA_CLIENTE not in df.columns:
        print(f"  AVISO: coluna '{COLUNA_CLIENTE}' não encontrada em '{nome_base}'. Pulando remoção.")
        return df
    antes = len(df)
    df = df[~df[COLUNA_CLIENTE].isin(CLIENTES_MATEUS)].reset_index(drop=True)
    removidas = antes - len(df)
    print(f"  Removidas {removidas} linha(s) dos clientes Mateus em '{nome_base}'")
    return df


def mesclar_base_mateus(df_normal, nome_mateus):
    """Carrega a base Mateus, remove linhas extras e appenda na base normal."""
    df_mateus = carregar_base(nome_mateus)
    df_mateus = remover_linhas_extras(df_mateus, nome_mateus)
    df_resultado = pd.concat([df_normal, df_mateus], ignore_index=True)
    print(f"  Mesclado '{nome_mateus}' -> total agora: {len(df_resultado)} linhas")
    return df_resultado


def salvar_base(df, nome_base):
    """Salva o DataFrame com todos os valores como texto puro."""
    arquivo = caminho(nome_base)

    wb = load_workbook(arquivo)
    ws = wb.active

    # Limpa todo o conteúdo existente
    ws.delete_rows(1, ws.max_row)

    # Escreve cabeçalho + dados forçando cada célula como string
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=str(value) if pd.notna(value) else "")
            cell.data_type = "s"

    wb.save(arquivo)
    print(f"  Salvo '{nome_base}.xlsx' com {len(df)} linhas\n")


def deletar_base_mateus(nome_mateus):
    """Apaga o arquivo Mateus após a mesclagem."""
    arquivo = caminho(nome_mateus)
    if arquivo.exists():
        arquivo.unlink()
        print(f"  Arquivo '{nome_mateus}.xlsx' deletado")


# ============================================================
# EXECUÇÃO PRINCIPAL
# ============================================================

def main():
    print("=" * 60)
    print("TRATAMENTO DAS BASES MTRIX")
    print("=" * 60)

    for nome_normal, nome_mateus in PARES_BASES.items():
        print(f"\n>>> Processando: '{nome_normal}'")

        # 1. Carrega a base normal
        df = carregar_base(nome_normal)

        # 2. Remove linhas extras (linhas em branco abaixo do cabeçalho)
        df = remover_linhas_extras(df, nome_normal)

        # 3. Remove os clientes Mateus da base normal
        df = remover_clientes_mateus(df, nome_normal)

        # 4. Carrega a base Mateus, remove linhas extras dela e mescla
        df = mesclar_base_mateus(df, nome_mateus)

        # 5. Sobrescreve a base normal com o resultado final
        salvar_base(df, nome_normal)

        # 6. Deleta o arquivo Mateus
        deletar_base_mateus(nome_mateus)

    print("=" * 60)
    print("Processo concluído! 4 bases finais salvas.")
    print("=" * 60)


if __name__ == "__main__":
    main()